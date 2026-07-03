"""지폭스케줄 결과를 PM23 생산지폭 표(.xlsx)로 출력."""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


def export_schedule_bytes(rows, total, *, jong, lot, cp, jeong_lens, choji=0):
    wb = Workbook()
    ws = wb.active
    ws.title = "지폭스케줄"

    thin = Side(style="thin", color="888888")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="CDE6F5")
    bold = Font(bold=True, size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    headers = ["계획 LOT", "23호기요청지폭", "초지 생산\n실 길이", "초지\n생산스플수",
               "비 고", "생산길이", "총생산길이"]
    ws.append(headers)
    for c in range(1, 8):
        cell = ws.cell(1, c)
        cell.font = bold
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = border

    start = 2
    for i, r in enumerate(rows):
        ri = start + i
        ws.cell(ri, 2, r.width)
        ws.cell(ri, 3, r.choji_real)
        ws.cell(ri, 4, r.spools)
        ws.cell(ri, 5, r.bigo())
        ws.cell(ri, 6, r.saengsan)
        for c in range(1, 8):
            ws.cell(ri, c).border = border
            ws.cell(ri, c).alignment = left if c == 5 else center
        ws.cell(ri, 3).number_format = "#,##0"
        ws.cell(ri, 6).number_format = "#,##0"

    last = start + len(rows) - 1 if rows else start
    # 계획 LOT 셀(병합): 지종/LOT/평량 + 정길이 목록
    jeong_text = "\n".join(f"{j:,}m" for j in sorted(set(jeong_lens)))
    lot_text = f"{jong}\n{lot}\n감열지 {cp}\n\n{jeong_text}"
    if rows:
        ws.merge_cells(start_row=start, start_column=1, end_row=last, end_column=1)
        ws.merge_cells(start_row=start, start_column=7, end_row=last, end_column=7)
    c1 = ws.cell(start, 1, lot_text)
    c1.alignment = center
    c1.font = Font(bold=True, size=10)
    c7 = ws.cell(start, 7, total)
    c7.alignment = center
    c7.number_format = "#,##0"
    c7.font = Font(bold=True, size=11)

    widths = [24, 10, 13, 11, 60, 13, 13]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()
