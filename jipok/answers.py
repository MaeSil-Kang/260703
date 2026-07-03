"""해답 Excel(학습\\Case+1~20+문제.xlsx, 파일명은 '문제'지만 내용은 해답) 로더.

각 Case = 헤더행(계획LOT/지종/CP) + 여러 PM23 생산지폭 행.
비고 문자열을 파싱해 (순서, 정길이, 회) 토큰 + 여유분 + 초출파지를 추출한다.
"""
import re
from dataclasses import dataclass, field
from openpyxl import load_workbook

BIGO_ORDER = re.compile(r"\(순서(\d+)번\)\s*([\d,]+)\s*m\s*용\s*\*\s*(\d+)\s*회")
BIGO_YEOYU_PLAIN = re.compile(r"여유분\s*(\d+)\s*회")
BIGO_YEOYU_LEN = re.compile(r"여유분\s*([\d,]+)\s*m\s*용?\s*(\d+)\s*회")
BIGO_CHOCHUL = re.compile(r"초출파지\s*([\d,]+)\s*m")


def _num(s):
    return int(str(s).replace(",", "").replace("m", "").strip())


@dataclass
class BigoToken:
    order: int          # 순서 (0 = 여유분/초출파지)
    jeong_len: int      # 정길이 (m); 초출파지는 길이
    count: int          # 회
    kind: str = "order"  # order | yeoyu | chochul


@dataclass
class AnswerRow:
    width: int                 # 원지지폭(생산폭)
    choji_real: int            # 초지생산실길이
    spools: int                # 초지생산스플수
    bigo: str
    saengsan: int              # 생산길이
    tokens: list = field(default_factory=list)

    @property
    def total_makki(self):
        return sum(t.count for t in self.tokens if t.kind in ("order", "yeoyu"))

    @property
    def jeong_lens(self):
        return sorted({t.jeong_len for t in self.tokens if t.kind == "order"})

    @property
    def yeoyu(self):
        return sum(t.count for t in self.tokens if t.kind == "yeoyu")


@dataclass
class AnswerCase:
    no: int
    jong: str
    cp: int
    teukgam: bool
    lot: str
    rows: list = field(default_factory=list)
    total: int = 0           # 총생산길이(해답)


def parse_bigo(bigo: str):
    if not bigo:
        return []
    toks = []
    for m in BIGO_ORDER.finditer(bigo):
        toks.append(BigoToken(int(m.group(1)), _num(m.group(2)), int(m.group(3)), "order"))
    for m in BIGO_YEOYU_LEN.finditer(bigo):
        toks.append(BigoToken(0, _num(m.group(1)), int(m.group(2)), "yeoyu"))
    # plain 여유분 N회 (정길이 명시 없음) — LEN 매칭과 겹치지 않게 처리
    bigo_wo_len = BIGO_YEOYU_LEN.sub("", bigo)
    for m in BIGO_YEOYU_PLAIN.finditer(bigo_wo_len):
        toks.append(BigoToken(0, 0, int(m.group(1)), "yeoyu"))
    for m in BIGO_CHOCHUL.finditer(bigo):
        toks.append(BigoToken(0, _num(m.group(1)), 0, "chochul"))
    return toks


def _parse_cp(jong: str):
    teukgam = ("SFH" in jong) or ("SL" in jong) or ("HL" in jong)
    m = re.search(r"(\d{2,3})", jong.split("\n")[-1])
    cp = int(m.group(1)) if m else 0
    return cp, teukgam


def load_answers(path: str):
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    cases = []
    cur = None
    for row in ws.iter_rows(values_only=True):
        # A열이 비어 있어 데이터는 B열(index 1)부터 시작
        c1 = row[1] if len(row) > 1 else None
        if c1 and isinstance(c1, str) and c1.strip().startswith("Case"):
            cur = AnswerCase(no=int(re.search(r"\d+", c1).group()), jong="", cp=0,
                             teukgam=False, lot="")
            cases.append(cur)
            _hdr_next = True
            continue
        if cur is None:
            continue
        # 헤더행("계획 LOT"라벨) 건너뜀
        if c1 and isinstance(c1, str) and "계획" in c1:
            continue
        width = row[3]
        if c1 and isinstance(c1, str) and "감열지" in c1 and not cur.jong:
            cur.jong = c1.strip()
            cur.cp, cur.teukgam = _parse_cp(c1)
            cur.lot = next((p for p in c1.split("\n") if p.strip().isdigit()), "")
        if isinstance(width, (int, float)):
            bigo = row[6] or ""
            r = AnswerRow(width=int(width),
                          choji_real=int(row[4]) if isinstance(row[4], (int, float)) else 0,
                          spools=int(row[5]) if isinstance(row[5], (int, float)) else 0,
                          bigo=str(bigo),
                          saengsan=int(row[14]) if isinstance(row[14], (int, float)) else 0,
                          tokens=parse_bigo(str(bigo)))
            cur.rows.append(r)
            if isinstance(row[15], (int, float)) and row[15]:
                cur.total = int(row[15])
    return cases


if __name__ == "__main__":
    import sys
    cases = load_answers(sys.argv[1])
    for c in cases:
        print(f"Case {c.no}: {c.jong.splitlines()[-1] if c.jong else '?'} CP{c.cp} "
              f"teukgam={c.teukgam} rows={len(c.rows)} total={c.total}")
